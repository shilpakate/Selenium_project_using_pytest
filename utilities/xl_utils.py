import openpyxl

xl_file_name = r"C:\Users\One\Desktop\nopcom3\test_data\proj_ddt.xlsx"


def get_rows_count(file_name=xl_file_name):
    workbook = openpyxl.load_workbook(file_name)
    active_sheet = workbook.active
    return active_sheet.max_row


def get_columns_count(file_name=xl_file_name):
    workbook = openpyxl.load_workbook(file_name)
    active_sheet = workbook.active
    return active_sheet.max_column


def read_data(row_num, col_num, file_name=xl_file_name):
    workbook = openpyxl.load_workbook(file_name)
    active_sheet = workbook.active
    return active_sheet.cell(row=row_num, column=col_num).value


def write_data(row_num, col_num, data, file_name=xl_file_name):
    workbook = openpyxl.load_workbook(file_name)
    active_sheet = workbook.active
    active_sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file_name)
